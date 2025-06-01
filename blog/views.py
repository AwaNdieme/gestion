from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from . import forms
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from .forms import BureauForm, ServiceForm, InfirmierForm, MedecinForm, PatientForm, RendezVousForm
from .models import Patient,Infirmier,RendezVous,Medecin,Bureau,Service
import csv
import openpyxl
from django.contrib import messages
import logging
from django.conf import settings


from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


from django.shortcuts import render
from .models import Patient, Medecin, Infirmier, Service


## fICHIER LOG 

from django.conf import settings
import os

def view_logs(request):
    """Vue pour afficher le contenu du fichier de logs"""
    # Chemin vers votre fichier de logs (ajustez selon votre configuration)
    log_file_path = os.path.join(settings.BASE_DIR, 'logs/app.log')
    
    try:
        with open(log_file_path, 'r') as f:
            log_content = f.read()
        
        # Si le fichier est trop grand, on ne prend que les 1000 dernières lignes
        log_lines = log_content.split('\n')
        if len(log_lines) > 1000:
            log_content = '\n'.join(log_lines[-1000:])
            message = "Affichage des 1000 dernières lignes :\n\n"
        else:
            message = "Contenu complet du fichier de logs :\n\n"
            
        return HttpResponse(f"<pre>{message}{log_content}</pre>")
    
    except FileNotFoundError:
        return HttpResponse("Fichier de logs non trouvé", status=404)
    except Exception as e:
        return HttpResponse(f"Erreur lors de la lecture du fichier : {str(e)}", status=500)

logger = logging.getLogger(__name__)







def admin_Dashboard(request):
    logger.info(f"Consultation du Dashboard par l'utilisateur {request.user.username}")
    context = {
        'total_patients': Patient.objects.count() or 0,
        'total_medecins': Medecin.objects.count() or 0,
        'total_infirmiers': Infirmier.objects.count() or 0,
        'total_services': Service.objects.count() or 0,
    }
    return render(request, 'blog/admin_dashboard.html', context)

##########################
from .forms import AdminProfileForm

@login_required
def admin_profile(request):
    logger.info(f"Consultation du profile par l'utilisateur {request.user.username}")
    user = request.user
    if request.method == 'POST':
        form = AdminProfileForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Profil mis à jour avec succès.")
            return redirect('admin-profile')
    else:
        form = AdminProfileForm(instance=user)

    return render(request, 'blog/admin-profile.html', {'form': form})


def login_page(request):
    logger.info(f"Tentative de connexion par l'utilisateur {request.POST.get('username', 'inconnu')}")
    form = forms.LoginForm()
    message = ''
    
    if request.method == 'POST':
        form = forms.LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('accueil')  # Redirection vers la page d'accueil
            else:
                message = 'Identifiants invalides,il vous reste 5 tentatives !!! '
                return render(request, 'blog/login.html', context={'form': form, 'message': message})
    
    return render(request, 'blog/login.html', context={'form': form, 'message': message})

@login_required
def accueil(request):
    logger.info(f"Consultation de l'accueil par l'utilisateur {request.user.username}")
    services = Service.objects.all()
    patients = Patient.objects.all()
    medecins = Medecin.objects.all()
    infirmiers = Infirmier.objects.all()

    context = {
        'services': services,
        'patients': patients,
        'medecins': medecins,
        'infirmiers': infirmiers,

         #  Ajout des compteurs numériques pour lincrementation
        'nb_services': services.count(),
        'nb_patients': patients.count(),
        'nb_medecins': medecins.count(),
        'nb_infirmiers': infirmiers.count(),
    }
    return render(request, 'blog/accueil.html', context)

def logout_view(request):
    logger.info(f"Déconnexion de l'utilisateur :inconnu")
    """Vue pour la déconnexion"""
    logout(request)
    return redirect('login')

from django.shortcuts import render, get_object_or_404, redirect
from .models import Service
from .forms import ServiceForm

# Vue principale de gestion des services
from django.shortcuts import render, redirect
from django.contrib import messages

def admin_Service(request):
    logger.info(f"Consultation des services par l'utilisateur {request.user.username}")
    services = Service.objects.all()
    recherche = request.GET.get('recherche', '')
    selected_service = request.GET.get('service', '')

    # Filtrer par nom
    if recherche:
        services = services.filter(nom_service__icontains=recherche)

    # Filtrer par ID
    if selected_service:
        services = services.filter(id=selected_service)

    # Gestion de l'ajout de service
    if request.method == 'POST':
        form = ServiceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Service ajouté avec succès!")
            return redirect('admin_Service')  # Redirige pour éviter les re-soumissions
    else:
        form = ServiceForm()  # Formulaire vide pour GET

    context = {
        'services': services,
        'selected_service': selected_service,
        'recherche': recherche,
        'form': form,
    }
    return render(request, 'blog/admin_Service.html', context)

# Ajouter un service (avec ServiceForm)

def ajouter_service(request):
    logger.info(f"Ajout d'un service par l'utilisateur {request.user.username}")
    if request.method == 'POST':
        form = ServiceForm(request.POST)
        if form.is_valid():
            form.save()  # Enregistrement automatique dans la BD
            return redirect('admin_Service')
    return redirect('admin_Service')
def modifier_service(request, pk):
    logger.info(f"Modification du service {pk} par l'utilisateur {request.user.username}")
    service = get_object_or_404(Service, pk=pk)  # Récupère le service à modifier
    
    if request.method == 'POST':
        form = ServiceForm(request.POST, instance=service)  # Lié au service existant
        if form.is_valid():
            form.save()  # Mise à jour
            messages.success(request, 'Service modifié avec succès!')
            return redirect('admin_Service')
        else:
            # Si le formulaire n'est pas valide, on l'affiche avec les erreurs
            messages.error(request, 'Erreur lors de la modification. Vérifiez les champs.')
    else:
        # GET: Affichage du formulaire pré-rempli
        form = ServiceForm(instance=service)
    
    # Rendu du template avec le formulaire (pré-rempli ou avec erreurs)
    return render(request, 'votre_template.html', {
        'form': form,
        'service': service,
        'action': 'Modifier'
    })

#  Supprimer un service
# ===============================
def supprimer_service(request, pk):
    logger.info(f"Suppression du service {pk} par l'utilisateur {request.user.username}")
    service = get_object_or_404(Service, pk=pk)
    if request.method == 'POST':
        service.delete()
    return redirect('admin_Service')



import csv
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import logging

logger = logging.getLogger(__name__)

@login_required
def export_services_csv(request):
    """Exporter les services au format CSV"""
    try:
        logger.info(f"Export des services au format CSV par l'utilisateur {request.user.username}")
        
        # Créer la réponse HTTP avec le bon type de contenu
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="services.csv"'
        
        # Ajouter le BOM UTF-8 pour Excel
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_ALL)
        
        # Écrire l'en-tête
        writer.writerow(['ID', 'Nom du Service', 'Description'])
        
        # Écrire les données des services
        services = Service.objects.all().order_by('id')
        for service in services:
            writer.writerow([
                service.id,
                service.nom_service or '',  # Gérer les valeurs nulles
                service.description or ''   # Gérer les valeurs nulles
            ])
        
        logger.info(f"Export CSV réussi - {services.count()} services exportés")
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export CSV: {str(e)}")
        # Retourner une réponse d'erreur appropriée
        return HttpResponse(
            "Erreur lors de l'export des services", 
            status=500,
            content_type='text/plain'
        )


# ===============================
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment

def export_services_excel(request):
    logger.info(f"Export des services au format Excel par l'utilisateur {request.user.username}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Services"

    # En-têtes stylés
    ws.append(['ID', 'Nom du Service', 'Description'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Données
    for s in Service.objects.all():
        ws.append([s.id, s.nom_service, s.description])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="services.xlsx"'
    wb.save(response)
    return response

# ===============================
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from .models import Service

def export_services_pdf(request):
    logger.info(f"Export des services au format PDF par l'utilisateur {request.user.username}")
    try:
        # Configuration du document PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="liste_services_{datetime.now().strftime("%Y%m%d")}.pdf"'

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=40, leftMargin=40,
            topMargin=60, bottomMargin=40
        )

        elements = []
        styles = getSampleStyleSheet()

        # Styles pour l'en-tête
        header_style = ParagraphStyle(
            'header_style',
            parent=styles['Normal'],
            fontSize=16,
            textColor=colors.black,
            alignment=1,  # 1 = centre
            spaceAfter=10,
            leading=15
        )
        
        subheader_style = ParagraphStyle(
            'subheader_style',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,
            spaceBefore=5,
            spaceAfter=10
        )

        # En-tête avec logos et texte
        header_data = [
            [Image('static/blog/img/uit.jpg', width=2*cm, height=2*cm), 
             [Paragraph("<b>HÔPITAL UNIVERSITAIRE DE L'UIT</b>", header_style),
              Paragraph("Centre Hospitalo-Universitaire de Référence", subheader_style)],
             Image('static/blog/img/dr.jpg', width=2*cm, height=2*cm)]
        ]
        
        header_table = Table(header_data, colWidths=[4*cm, 10*cm, 4*cm])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(header_table)

        # Informations de l'hôpital
        elements.append(Paragraph("ANNEE 2024-2025", styles['Heading2']))
        elements.append(Spacer(1, 12))

        hospital_info = [
            Paragraph("123 Avenue de la Santé", styles['Normal']),
            Paragraph("Ville: Thies", styles['Normal']),
            Paragraph("Tél: +221 33 945 67 89", styles['Normal']),
            Paragraph("Email: contact@hopital-uit.sn", styles['Normal']),
            Spacer(1, 24),
            Paragraph("<b>LISTE DES SERVICES MÉDICAUX</b>", styles['Title']),
            Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
            Spacer(1, 24)
        ]
        elements.extend(hospital_info)

        # Tableau des services
        data = [['ID', 'Nom du Service', 'Description']]
        services = Service.objects.all().order_by('nom_service')
        
        for service in services:
            data.append([
                service.id,
                service.nom_service,
                service.description or "Aucune description disponible"  # Gestion des descriptions vides
            ])

        # Création du tableau avec largeurs de colonnes proportionnelles
        table = Table(data, colWidths=[2*cm, 5*cm, 9*cm], repeatRows=1)
        
        # Style du tableau - adapté mais cohérent avec la charte
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E5F8C')),  # Bleu hospitalier
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Centrer uniquement la colonne ID
            ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F8FF')),  # Fond bleu clair
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 24))

        # Pied de page
        footer_style = ParagraphStyle(
            'footer_style',
            fontSize=8,
            textColor=colors.HexColor('#6c757d'),
            alignment=1  # Centré
        )
        
        elements.append(Paragraph(
            f"Total services: {services.count()} • © {datetime.now().year} Hôpital Universitaire de l'UIT",
            footer_style
        ))

        # Génération du PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du PDF: {str(e)}", status=500)
# ===============================
# Fin de la gestion des services

#############  BUREAU #############from django.db.models import Q

def admin_Bureau(request):
    logger.info(f"Consultation des bureaux par l'utilisateur {request.user.username}")
    bureaux = Bureau.objects.select_related('service').all().order_by('etage', 'numero_bureau')
    services = Service.objects.all()
    
    # Récupération des paramètres de filtrage
    recherche = request.GET.get('recherche', '')
    selected_service = request.GET.get('service', '')
    etage = request.GET.get('etage', '')
    
    # Application des filtres
    if recherche:
        bureaux = bureaux.filter(
            Q(numero_bureau__icontains=recherche) |
            Q(service__nom_service__icontains=recherche)
        )
    
    if selected_service:
        bureaux = bureaux.filter(service__id=selected_service)
    
    if etage:
        bureaux = bureaux.filter(etage=etage)
    
    # Récupération des étages distincts
    etages = Bureau.objects.values_list('etage', flat=True).distinct().order_by('etage')
    
    context = {
        'bureaux': bureaux,
        'services': services,
        'etages': etages,
        'recherche': recherche,
        'selected_service': selected_service,
        'selected_etage': etage,
        'form': BureauForm()
    }
    return render(request, 'blog/admin_Bureau.html', context)


def ajouter_bureau(request):
    logger.info(f"Ajout d'un bureau par l'utilisateur {request.user.username}")
    if request.method == 'POST':
        form = BureauForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_Bureau')
    return redirect('admin_Bureau')



def modifier_bureau(request, pk):
    logger.info(f"Modification du bureau {pk} par l'utilisateur {request.user.username}")
    bureau = get_object_or_404(Bureau, pk=pk)
    if request.method == 'POST':
        form = BureauForm(request.POST, instance=bureau)
        if form.is_valid():
            form.save()
            return redirect('admin_Bureau')
    return redirect('admin_Bureau')


def supprimer_bureau(request, pk):
    logger.info(f"Suppression du bureau {pk} par l'utilisateur {request.user.username}")
    bureau = get_object_or_404(Bureau, pk=pk)
    if request.method == 'POST':
        bureau.delete()
    return redirect('admin_Bureau')


#############################

def export_bureau_csv(request):
    logger.info(f"Export des bureaux au format CSV par l'utilisateur {request.user.username}")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bureaux.csv"'
    writer = csv.writer(response)
    writer.writerow(['Numero Bureau', 'Etage', 'Service'])

    for b in Bureau.objects.all():
        writer.writerow([b.numero_bureau, b.etage, b.service.nom_service])
    return response


def export_bureau_excel(request):
    logger.info(f"Export des bureaux au format Excel par l'utilisateur {request.user.username}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bureaux"
    ws.append(['Numero Bureau', 'Etage', 'Service'])

    for b in Bureau.objects.all():
        ws.append([b.numero_bureau, b.etage, b.service.nom_service])

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="bureaux.xlsx"'
    wb.save(response)
    return response
############################
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from .models import Bureau

def export_bureau_pdf(request):
    logger.info(f"Export des bureaux au format PDF par l'utilisateur {request.user.username}")
    try:
        # Configuration du document PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="liste_bureaux_{datetime.now().strftime("%Y%m%d")}.pdf"'

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=40, leftMargin=40,
            topMargin=60, bottomMargin=40
        )

        elements = []
        styles = getSampleStyleSheet()

        # Styles pour l'en-tête
        header_style = ParagraphStyle(
            'header_style',
            parent=styles['Normal'],
            fontSize=16,
            textColor=colors.black,
            alignment=1,  # 1 = centre
            spaceAfter=10,
            leading=13
        )
        
        subheader_style = ParagraphStyle(
            'subheader_style',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,
            spaceBefore=5,
            spaceAfter=10
        )

        # En-tête avec logos et texte
        header_data = [
            [Image('static/blog/img/uit.jpg', width=2*cm, height=2*cm), 
             [Paragraph("<b>HÔPITAL UNIVERSITAIRE DE L'UIT</b>", header_style),
              Paragraph("Centre Hospitalo-Universitaire de Référence", subheader_style)],
             Image('static/blog/img/dr.jpg', width=2*cm, height=2*cm)]
        ]
        
        header_table = Table(header_data, colWidths=[4*cm, 10*cm, 4*cm])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(header_table)

        # Informations de l'hôpital
        elements.append(Paragraph("ANNEE 2024-2025", styles['Heading2']))
        elements.append(Spacer(1, 12))

        hospital_info = [
            Paragraph("123 Avenue de la Santé", styles['Normal']),
            Paragraph("Ville: Thies", styles['Normal']),
            Paragraph("Tél: +221 33 945 67 89", styles['Normal']),
            Paragraph("Email: contact@hopital-uit.sn", styles['Normal']),
            Spacer(1, 24),
            Paragraph("<b>LISTE DES BUREAUX - SERVICE LOGISTIQUE</b>", styles['Title']),
            Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
            Spacer(1, 24)
        ]
        elements.extend(hospital_info)

        # Tableau des bureaux
        data = [['Numéro', 'Étage', 'Service']]
        bureaux = Bureau.objects.select_related('service').order_by('etage', 'numero_bureau')
        
        for bureau in bureaux:
            data.append([
                bureau.numero_bureau,
                f"Étage {bureau.etage}",
                bureau.service.nom_service if bureau.service else "Non attribué"
            ])

        # Création du tableau avec largeurs de colonnes proportionnelles
        table = Table(data, colWidths=[3*cm, 3*cm, 10*cm], repeatRows=1)
        
        # Style du tableau - adapté mais cohérent avec la charte
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E5F8C')),  # Bleu hospitalier
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F8FF')),  # Fond bleu clair
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 24))

        # Pied de page
        footer_style = ParagraphStyle(
            'footer_style',
            fontSize=8,
            textColor=colors.HexColor('#6c757d'),
            alignment=1  # Centré
        )
        
        elements.append(Paragraph(
            f"Total bureaux: {bureaux.count()} • © {datetime.now().year} Hôpital Universitaire de l'UIT",
            footer_style
        ))

        # Génération du PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du PDF: {str(e)}", status=500)

#############   INFIRMIER ##############
#Affichage des infirmiers
from django.db.models import Q

def admin_Infirmier(request):
    logger.info(f"Consultation des infirmiers par l'utilisateur {request.user.username}")
    infirmiers = Infirmier.objects.select_related('service').all()
    recherche = request.GET.get('recherche', '')
    selected_grade = request.GET.get('grade', '')
    selected_service = request.GET.get('service', '')

    # Filtres combinés
    if recherche:
        infirmiers = infirmiers.filter(
            Q(nom__icontains=recherche) | 
            Q(prenom__icontains=recherche)
    )
    
    if selected_grade:
        infirmiers = infirmiers.filter(grade__iexact=selected_grade)
        
    if selected_service:
        infirmiers = infirmiers.filter(service__id=selected_service)

    # Récupération des grades distincts existants
    grades = Infirmier.objects.values_list('grade', flat=True).distinct()

    context = {
        'infirmiers': infirmiers,
        'grades': grades,
        'recherche': recherche,
        'selected_grade': selected_grade,
        'selected_service': selected_service,
        'form': InfirmierForm(),
        'services': Service.objects.all()
    }
    return render(request, 'blog/admin_Infirmier.html', context)

#  Ajout d’un infirmier via InfirmierForm
def ajouter_infirmier(request):
    logger.info(f"Ajout d'un infirmier par l'utilisateur {request.user.username}")
    if request.method == "POST":
        form = InfirmierForm(request.POST)
        if form.is_valid():
            form.save()
    return redirect('admin_Infirmier')

#  Modifier infirmier
def modifier_infirmier(request, pk):
    logger.info(f"Modification de l'infirmier {pk} par l'utilisateur {request.user.username}")
    infirmier = get_object_or_404(Infirmier, pk=pk)
    if request.method == "POST":
        form = InfirmierForm(request.POST, instance=infirmier)
        if form.is_valid():
            form.save()
            return redirect('admin_Infirmier')
    return redirect('admin_Infirmier')

#  Supprimer infirmier
def supprimer_infirmier(request, pk):
    logger.info(f"Suppression de l'infirmier {pk} par l'utilisateur {request.user.username}")
    infirmier = get_object_or_404(Infirmier, pk=pk)
    if request.method == "POST":
        infirmier.delete()
    return redirect('admin_Infirmier')




def export_infirmiers_csv(request):
    logger.info(f"Export des infirmiers au format CSV par l'utilisateur {request.user.username}")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="infirmiers.csv"'
    writer = csv.writer(response)
    writer.writerow(['Nom', 'Prénom', 'Grade', 'Téléphone', 'Service'])

    for i in Infirmier.objects.all():
        writer.writerow([i.nom, i.prenom, i.grade, i.telephone, i.service.nom_service if i.service else ''])
    
    return response


def export_infirmiers_excel(request):
    logger.info(f"Export des infirmiers au format Excel par l'utilisateur {request.user.username}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Infirmiers"
    ws.append(['Nom', 'Prénom', 'Grade', 'Téléphone', 'Service'])

    for i in Infirmier.objects.all():
        ws.append([i.nom, i.prenom, i.grade, i.telephone, i.service.nom_service if i.service else ''])

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="infirmiers.xlsx"'
    wb.save(response)
    return response

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime

def export_infirmiers_pdf(request):
    logger.info(f"Export des infirmiers au format PDF par l'utilisateur {request.user.username}")
    # Configuration du document
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                          rightMargin=40, leftMargin=40,
                          topMargin=60, bottomMargin=60)
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Style pour l'en-tête des infirmiers (couleur verte médicale)
    infirmier_header_style = ParagraphStyle(
        'infirmier_header',
        parent=styles['Heading1'],
        textColor=colors.HexColor('#2e7d32'),  # Vert médical
        spaceAfter=12,
        alignment=1
    )
    
    # 1. En-tête avec style spécifique aux infirmiers
    elements.append(Paragraph("<img src='static/blog/img/infirmier.jpg' width='50' height='50'/><br/>", 
                           styles['Normal']))
    elements.append(Paragraph("LISTE DU PERSONNEL INFIRMIER UIT", infirmier_header_style))
    
    # Informations complémentaires
    info_style = ParagraphStyle(
        'info_style',
        parent=styles['Normal'],
        textColor=colors.HexColor('#1b5e20'),
        fontSize=10,
        leading=12
    )
    
    elements.extend([
        Paragraph(f"<b>Hôpital Général de l'UIT</b>", info_style),
        
        Paragraph(f"Service des Soins Infirmiers", info_style),
       
        Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}", info_style),
        Spacer(1, 24)
    ])

    # 2. Données des infirmiers - Style différent des médecins
    data = [
        ['ID', 'Nom', 'Prénom', 'Grade', 'Téléphone', 'Service']
    ]
    
    infirmiers = Infirmier.objects.select_related('service').order_by('nom', 'prenom')
    
    for idx, infirmier in enumerate(infirmiers, start=1):
        data.append([
            str(idx),
            infirmier.nom.upper(),
            infirmier.prenom,
            Paragraph(infirmier.grade, styles['Normal']),
            infirmier.telephone or "N/A",
            infirmier.service.nom_service if infirmier.service else "Non affecté"
        ])

    # 3. Style du tableau spécifique aux infirmiers
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#81c784')),  # Vert clair
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#e8f5e9')),  # Vert très clair
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#a5d6a7')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 24))
    
    # 4. Pied de page spécifique
    footer_style = ParagraphStyle(
        'footer_style',
        parent=styles['Italic'],
        textColor=colors.HexColor('#2e7d32'),
        fontSize=8,
        alignment=1
    )
    
    elements.append(Paragraph(
        "Service des Soins Infirmiers | © Hôpital Général {}".format(datetime.now().year),
        footer_style
    ))

    # Génération du PDF
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="liste_infirmiers_{}.pdf"'.format(
        datetime.now().strftime('%Y%m%d')
    )
    return response
###############################  FIN INFIRMIER ###########


#############   MEDECIN ##############from django.shortcuts import render, redirect
from .forms import MedecinForm
from .models import Medecin, Service, Bureau
from django.db.models import Q

def admin_medecin(request):
    logger.info(f"Consultation des médecins par l'utilisateur {request.user.username}")
    # Traitement du formulaire s’il y a une soumission POST
    if request.method == 'POST':
        form = MedecinForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_medecin') 
    else:
        form = MedecinForm()

    # Gestion des filtres de recherche (GET)
    recherche = request.GET.get('recherche', '')
    specialite = request.GET.get('specialite', '')
    service = request.GET.get('service', '')

    medecins = Medecin.objects.all()

    if recherche:
        medecins = medecins.filter(
            Q(nom__icontains=recherche) |
            Q(prenom__icontains=recherche)
        )
    if specialite:
        medecins = medecins.filter(specialite__icontains=specialite)
    if service:
        medecins = medecins.filter(service__id=service)

    context = {
        'medecins': medecins,
        'recherche': recherche,
        'selected_specialite': specialite,
        'selected_service': service,
        'form': form,
        'services': Service.objects.all(),
        'bureaux': Bureau.objects.all(),
        'specialites': Medecin.objects.values_list('specialite', flat=True).distinct(),
    }
    return render(request, 'blog/admin_Medecin.html', context)

def ajouter_medecin(request):
    logger.info(f"Ajout d'un médecin par l'utilisateur {request.user.username}")
    if request.method == 'POST':
        form = MedecinForm(request.POST)
        if form.is_valid():
            form.save()
    return redirect('admin_Medecin')




def modifier_medecin(request, pk):
    logger.info(f"Modification du médecin {pk} par l'utilisateur {request.user.username}")
    medecin = get_object_or_404(Medecin, pk=pk)

    if request.method == 'POST':
        medecin.nom = request.POST.get('nom')
        medecin.prenom = request.POST.get('prenom')
        medecin.specialite = request.POST.get('specialite')
        medecin.email = request.POST.get('email')

        # Récupération des objets Service et Bureau à partir des IDs
        service_id = request.POST.get('service')
        bureau_id = request.POST.get('bureau')

        medecin.service = Service.objects.get(id=service_id) if service_id else None
        medecin.bureau = Bureau.objects.get(id=bureau_id) if bureau_id else None

        medecin.save()
        return redirect('admin_Medecin')




def supprimer_medecin(request, pk):
    logger.info(f"Suppression du médecin {pk} par l'utilisateur {request.user.username}")
    medecin = get_object_or_404(Medecin, pk=pk)
    if request.method == 'POST':
        medecin.delete()
    return redirect('admin_Medecin')



########## Exporter les médecins au format pdf, excel et csv ##########from io import BytesIO
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from datetime import datetime
from .models import Medecin

def export_medecin_pdf(request):
    logger.info(f"Export des médecins au format PDF par l'utilisateur {request.user.username}")
    # Configuration du document PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="liste_medecins_{datetime.now().strftime("%Y%m%d")}.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=72, leftMargin=72,
        topMargin=72, bottomMargin=72
    )

    elements = []
    styles = getSampleStyleSheet()

    # Styles pour l'en-tête
    header_style = ParagraphStyle(
        'header_style',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.black,
        alignment=1,  # 1 = centre
        spaceAfter=10,
        leading=13
    )
    
    subheader_style = ParagraphStyle(
        'subheader_style',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.black,
        alignment=1,
        spaceBefore=5,
        spaceAfter=10
    )

    # En-tête avec logos et texte
    header_data = [
        [Image('static/blog/img/uit.jpg', width=2*cm, height=2*cm), 
         [Paragraph("<b>HÔPITAL UNIVERSITAIRE DE L'UIT</b>", header_style),
          Paragraph("Centre Hospitalo-Universitaire de Référence", subheader_style)],
         Image('static/blog/img/dr.jpg', width=2*cm, height=2*cm)]
    ]
    
    header_table = Table(header_data, colWidths=[4*cm, 10*cm, 4*cm])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(header_table)

    # Informations de l'hôpital
    elements.append(Spacer(1, 20))
    hospital_info = [
        Paragraph("123 Avenue de la Santé", styles['Normal']),
        Paragraph("Ville: Thies", styles['Normal']),
        Paragraph("Tél: +221 33 945 67 89", styles['Normal']),
        Paragraph("Email: contact@hopital-uit.sn", styles['Normal']),
        Spacer(1, 12),
        Paragraph("<b>LISTE DES MÉDECINS</b>", styles['Title']),
        Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
        Spacer(1, 24)
    ]
    elements.extend(hospital_info)

    # Données des médecins - Même structure que votre version initiale
    data = [['Nom', 'Prénom', 'Spécialité', 'Email', 'Service', 'Bureau']]
    
    medecins = Medecin.objects.select_related('service', 'bureau').order_by('nom', 'prenom')
    
    for medecin in medecins:
        data.append([
            medecin.nom.upper(),
            medecin.prenom,
            medecin.specialite,
            medecin.email or "N/A",
            str(medecin.service) if medecin.service else "N/A",
            str(medecin.bureau) if medecin.bureau else "N/A"
        ])

    # Création du tableau - Mêmes dimensions que votre version initiale
    table = Table(data, repeatRows=1)  # Pas de colWidths spécifiées comme dans l'original
    
    # Style du tableau - Identique à votre version initiale
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005b96')),  # Même bleu foncé
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Tout centré comme original
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),  # Même taille de police
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Même fond beige
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),  # Même grille
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 24))
    
    # Pied de page - Style similaire à l'original
    footer = Paragraph(
        f"<i>Document confidentiel - © Hôpital Universitaire de l'UIT {datetime.now().year}</i>",
        ParagraphStyle(
            'footer',
            parent=styles['Italic'],
            fontSize=8,
            textColor=colors.HexColor('#6c757d'),
            alignment=1  # Centré
        )
    )
    elements.append(footer)

    # Génération du PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def export_medecin_excel(request):
    logger.info(f"Export des médecins au format Excel par l'utilisateur {request.user.username}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Médecins"
    ws.append(['Nom', 'Prénom', 'Spécialité', 'Email', 'Service', 'Bureau'])

    for m in Medecin.objects.all():
        ws.append([m.nom, m.prenom, m.specialite, m.email, str(m.service), str(m.bureau)])

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="medecins.xlsx"'
    wb.save(response)
    return response


def export_medecin_csv(request):
    logger.info(f"Export des médecins au format CSV par l'utilisateur {request.user.username}")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="medecins.csv"'

    writer = csv.writer(response)
    writer.writerow(['Nom', 'Prénom', 'Spécialité', 'Email', 'Service', 'Bureau'])
    for m in Medecin.objects.all():
        writer.writerow([m.nom, m.prenom, m.specialite, m.email, m.service, m.bureau])
    return response


###############################  FIN MEDECIN ###########
#############   PATIENT ##############


from django.db.models import Q

def admin_Patient(request):
    logger.info(f"Consultation des patients par l'utilisateur {request.user.username}")
    patients = Patient.objects.select_related('service', 'medecin').all()
    services = Service.objects.all()
    medecins = Medecin.objects.all()

    # Récupération des paramètres
    recherche = request.GET.get('recherche', '')
    selected_service = request.GET.get('service', '')
    selected_medecin = request.GET.get('medecin', '')

    # Application des filtres
    if recherche:
        patients = patients.filter(
            Q(nom__icontains=recherche) | 
            Q(prenom__icontains=recherche)
        )

    if selected_service:
        patients = patients.filter(service__id=selected_service)

    if selected_medecin:
        patients = patients.filter(medecin__id=selected_medecin)

    context = {
        'patients': patients,
        'services': services,
        'medecins': medecins,
        'recherche': recherche,
        'selected_service': selected_service,
        'selected_medecin': selected_medecin
    }
    return render(request, 'blog/admin_Patient.html', context)




def ajouter_patient(request):
    logger.info(f"Ajout d'un patient par l'utilisateur {request.user.username}")
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Patient ajouté avec succès.")
    return redirect('admin_Patient')


def modifier_patient(request, pk):
    logger.info(f"Modification du patient {pk} par l'utilisateur {request.user.username}")
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            messages.warning(request, "✏️ Patient modifié avec succès.")
    return redirect('admin_Patient')


def supprimer_patient(request, pk):
    logger.info(f"Suppression du patient {pk} par l'utilisateur {request.user.username}")
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        patient.delete()
        messages.error(request, "🗑️ Patient supprimé avec succès.")
    return redirect('admin_Patient')




import csv
from django.http import HttpResponse
from .models import Patient

def export_patients_csv(request):
    logger.info(f"Export des patients au format CSV par l'utilisateur {request.user.username}")
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="patients.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Nom', 'Prénom', 'Date de naissance', 'Sexe', 'Adresse', 'Service', 'Médecin'])

    for patient in Patient.objects.all():
        writer.writerow([
            patient.id,
            patient.nom,
            patient.prenom,
            patient.date_naissance,
            patient.sexe,
            patient.adresse,
            patient.service.nom_service if patient.service else '',
            f"{patient.medecin.prenom} {patient.medecin.nom}" if patient.medecin else ''
        ])
    
    return response
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse

def export_patients_excel(request):
    logger.info(f"Export des patients au format Excel par l'utilisateur {request.user.username}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"

    # En-têtes avec style
    headers = ['ID', 'Nom', 'Prénom', 'Date de naissance', 'Sexe', 'Adresse', 'Service', 'Médecin']
    header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    header_font = Font(bold=True)

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font

    # Lignes de données
    for p in Patient.objects.all():
        ws.append([
            p.id,
            p.nom,
            p.prenom,
            p.date_naissance,
            p.sexe,
            p.adresse,
            p.service.nom_service if p.service else '',
            f"{p.medecin.prenom} {p.medecin.nom}" if p.medecin else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="patients.xlsx"'
    wb.save(response)
    return response
###################from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from datetime import datetime
from django.http import HttpResponse
from .models import Patient

def export_patients_pdf(request):
    logger.info(f"Export des patients au format PDF par l'utilisateur {request.user.username}")
    # Configuration du document PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="liste_patients.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=40, leftMargin=40,
        topMargin=60, bottomMargin=40
    )

    elements = []
    styles = getSampleStyleSheet()

    # Styles pour l'en-tête
    header_style = ParagraphStyle(
        'header_style',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.black,
        alignment=1,  # 1 = centre
        spaceAfter=10,
        leading=13
    )
    
    subheader_style = ParagraphStyle(
        'subheader_style',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.black,
        alignment=1,
        spaceBefore=5,
        spaceAfter=10
    )

    # En-tête avec logos et texte
    header_data = [
        [Image('static/blog/img/uit.jpg', width=2*cm, height=2*cm), 
         [Paragraph("<b>HÔPITAL UNIVERSITAIRE DE L'UIT</b>", header_style),
          Paragraph("Centre Hospitalo-Universitaire de Référence", subheader_style)],
         Image('static/blog/img/dr.jpg', width=2*cm, height=2*cm)]
    ]
    
    header_table = Table(header_data, colWidths=[4*cm, 10*cm, 4*cm])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(header_table)

    # Informations de l'hôpital
    elements.append(Paragraph("ANNEE 2024-2025", styles['Heading2']))
    elements.append(Spacer(1, 12))

    hospital_info = [
        Paragraph("123 Avenue de la Santé", styles['Normal']),
        Paragraph("Ville: Thies", styles['Normal']),
        Paragraph("Tél: +221 33 945 67 89", styles['Normal']),
        Paragraph("Email: contact@hopital-uit.sn", styles['Normal']),
        Spacer(1, 24),
        Paragraph("<b>LISTE DES PATIENTS ENREGISTRÉS</b>", styles['Title']),
        Spacer(1, 12)
    ]
    elements.extend(hospital_info)
    
    # Données du tableau
    data = [['ID', 'Nom', 'Prénom', 'Naissance', 'Sexe', 'Adresse']]
    patients = Patient.objects.all().order_by('nom', 'prenom')

    for p in patients:
        data.append([
            p.id,
            p.nom.upper(),
            p.prenom,
            p.date_naissance.strftime('%d/%m/%Y') if p.date_naissance else '',
            p.sexe,
            (p.adresse[:35] + '...') if len(p.adresse) > 35 else p.adresse
        ])

    # Création du tableau
    table = Table(data, colWidths=[0.8*cm, 2.5*cm, 2.5*cm, 2*cm, 1.5*cm, 5*cm], repeatRows=1)

    # Style du tableau
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E5F8C')),  # Bleu hospitalier
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (3, -1), 'CENTER'),
        ('ALIGN', (4, 0), (4, -1), 'CENTER'),  # Centrer la colonne sexe
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F8FF'))  # Fond bleu clair
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))

    # Date de génération et pied de page
    footer_style = ParagraphStyle(
        'footer_style',
        fontSize=8,
        textColor=colors.HexColor('#6c757d'),
        alignment=2
    )
    elements.append(Paragraph(
        f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        footer_style
    ))

    elements.append(Paragraph(
        f"Total : {patients.count()} patients — © {datetime.now().year} Hôpital Universitaire de l'UIT",
        ParagraphStyle(
            'footer_center',
            fontSize=8,
            textColor=colors.HexColor('#6c757d'),
            alignment=1
        )
    ))

    # Générer le PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response
############  RV #########
from .models import RendezVous
from .forms import RendezVousForm


def admin_RendezVous(request):
    logger.info(f"Consultation des rendez-vous par l'utilisateur {request.user.username}")
    rendezVouss = RendezVous.objects.all()
    statut_filter = request.GET.get('statut', '')

    # Filtrage uniquement par statut
    if statut_filter:
        rendezVouss = rendezVouss.filter(statut=statut_filter)

    context = {
        'rendezVouss': rendezVouss,
        'statut_filter': statut_filter,
        'patients': Patient.objects.all(),
        'medecins':Medecin.objects.all()
    
    }
    return render(request, 'blog/admin_RendezVous.html', context)
    # Filtrage par nom du patient
    if recherche:
        rendezVouss = rendezVouss.filter(patient__nom__icontains=recherche)

    # Filtrage par ID
    if selected_rendezVous:
        rendezVouss = rendezVouss.filter(id=selected_rendezVous)

    form = RendezVousForm()  # Formulaire vide pour l’ajout

    context = {
        'rendezVouss': rendezVouss,
        'selected_rendezVous': selected_rendezVous,
        'recherche': recherche,
        'form': form,
        'patients': Patient.objects.all(),
        'medecins': Medecin.objects.all(),

    }
    return render(request, 'blog/admin_RendezVous.html', context)

#  Ajouter un rendez-vous
def ajouter_rendezVous(request):
    logger.info(f"Ajout d'un rendez-vous par l'utilisateur {request.user.username}")
    if request.method == "POST":
        form = RendezVousForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('admin_RendezVous')
    

#  Modifier un rendez-vous
def modifier_rendezVous(request, pk):
    logger.info(f"Modification du rendez-vous {pk} par l'utilisateur {request.user.username}")
    rendezVous = get_object_or_404(RendezVous, pk=pk)
    if request.method == 'POST':
        form = RendezVousForm(request.POST, instance=rendezVous)
        if form.is_valid():
            form.save()
        return redirect('admin_RendezVous')
    

#  Supprimer un rendez-vous
def supprimer_rendezVous(request, pk):
    logger.info(f"Suppression du rendez-vous {pk} par l'utilisateur {request.user.username}")
    rendezVous = get_object_or_404(RendezVous, pk=pk)
    if request.method == 'POST':
        rendezVous.delete()
    return redirect('admin_RendezVous')
# ===============================


########export rendezvous
import csv
from django.http import HttpResponse
from .models import RendezVous

def export_rendezvous_csv(request):
    logger.info(f"Export des rendez-vous au format CSV par l'utilisateur {request.user.username}")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rendezvous.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Statut', 'Patient', 'Médecin'])

    for rdv in RendezVous.objects.all():
        writer.writerow([
            rdv.date_rdv,
            rdv.statut,
            f"{rdv.patient.prenom} {rdv.patient.nom}",
            f"{rdv.medecin.prenom} {rdv.medecin.nom}" if rdv.medecin else "Non assigné"
        ])

    return response


import openpyxl
from django.http import HttpResponse
from .models import RendezVous
from openpyxl.styles import Font, Alignment, PatternFill



##################

def export_rendezvous_excel(request):
    logger.info(f"Export des rendez-vous au format Excel par l'utilisateur {request.user.username}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendez-vous"

    # En-têtes stylés
    headers = ['Date', 'Statut', 'Patient', 'Médecin']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")

    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # Données
    for rdv in RendezVous.objects.all():
        ws.append([
            str(rdv.date_rdv),
            rdv.statut,
            f"{rdv.patient.prenom} {rdv.patient.nom}",
            f"{rdv.medecin.prenom} {rdv.medecin.nom}" if rdv.medecin else "Non assigné"
        ])

    # Réponse
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="rendezvous.xlsx"'
    wb.save(response)
    return response
########################
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from io import BytesIO
from django.http import HttpResponse
from reportlab.pdfgen import canvas

def export_rendezvous_pdf(request):
    logger.info(f"Export des rendez-vous au format PDF par l'utilisateur {request.user.username}")
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rendezvous.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Créer un style pour le titre de l'en-tête

    header_style = ParagraphStyle(
        'header_style',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.black,
        alignment=1,  # 1 = centre
        spaceAfter=10,
        leading=13  # 
    )  
    subheader_style = ParagraphStyle(
    'subheader_style',
    parent=styles['Normal'],
    fontSize=12,
    textColor=colors.black,
    alignment=1,
    spaceBefore=5,  # Espace avant ce paragraphe
    spaceAfter=10
)
    # Créer un tableau pour l'en-tête avec les images et le texte
    from reportlab.platypus import Table, TableStyle
    header_data = [
        [Image('static/blog/img/uit.jpg', width=2*cm, height=2*cm), 
     [Paragraph("<b>HÔPITAL UNIVERSITAIRE DE L'UIT</b>", header_style),
      Paragraph("Le Bien-être étudiant, au cœur de Nos Priorités", subheader_style)],
         Image('static/blog/img/dr.jpg', width=2*cm, height=2*cm)]
    ]
    header_table = Table(header_data, colWidths=[4*cm, 10*cm, 4*cm])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(header_table)

    # Ajouter l'année universitaire
    elements.append(Paragraph("ANNEE 2024-2025", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # Ajouter les informations de l'hôpital
    hospital_info = [
        Paragraph("123 Avenue de la Santé", styles['Normal']),
        Paragraph("Ville: Thies", styles['Normal']),
        Paragraph("Tél: +33 345 67 78", styles['Normal']),
        Paragraph("Email: uit@univ_hopital.com", styles['Normal']),
        Spacer(1, 24),
        Paragraph("<b>Liste des Rendez-vous</b>", styles['Title']),
        Spacer(1, 12)
    ]
    elements.extend(hospital_info)

    # Créer un tableau pour les rendez-vous
    from reportlab.platypus import Table, TableStyle
    rendezvous_data = [["Date", "Statut", "Patient", "Médecin"]]
    
    for rdv in RendezVous.objects.all():
        rendezvous_data.append([
            str(rdv.date_rdv),
            rdv.statut,
            f"{rdv.patient.prenom} {rdv.patient.nom}",
            f"{rdv.medecin.prenom} {rdv.medecin.nom}" if rdv.medecin else "N/A"
        ])

    rendezvous_table = Table(rendezvous_data, colWidths=[4*cm, 3*cm, 5*cm, 5*cm])
    rendezvous_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(rendezvous_table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response
